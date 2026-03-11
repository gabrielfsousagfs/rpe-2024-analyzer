"""
Analisador de Inventários de GEE — Compensação por Créditos de Carbono
=======================================================================
Analisa PDFs do Registro Público de Emissões (RPE 2024) para identificar
empresas que compram créditos de carbono para compensação de emissões.

Seções analisadas:
    4.3 — Descrição de estratégias e projetos para gestão de emissões de GEE
    5.1 — Compensação de emissões
    5.2 — Reduções de emissões
"""

import os
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime

try:
    import fitz  # PyMuPDF
except ImportError:
    print("❌ PyMuPDF não encontrado. Instale com: pip install pymupdf")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
    TQDM_DISPONIVEL = True
except ImportError:
    TQDM_DISPONIVEL = False

# ── Configuração ──────────────────────────────────────────────────────────────
PASTA_PDFS       = os.environ.get("PASTA_PDFS", "./pdfs")
TAMANHO_CONTEXTO = 350   # caracteres ao redor do termo encontrado

_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\uFFFE\uFFFF]")

def sanitizar(texto: str) -> str:
    if not isinstance(texto, str):
        return str(texto) if texto is not None else ""
    limpo = _ILLEGAL_CHARS_RE.sub(" ", texto)
    return re.sub(r" {2,}", " ", limpo).strip()


# ── Padrões de detecção ───────────────────────────────────────────────────────
#
# CAMADA 1: Termos que confirmam diretamente compra/uso de créditos externos.
# Qualquer match aqui → SIM.
#
# CAMADA 2: Termos de compensação/neutralização genéricos. Só confirmam → SIM
# se houver pelo menos um match da Camada 1 no mesmo PDF.
#
# A separação evita dois tipos de erro:
#   (a) Falso positivo: empresa que apenas "reduz internamente" mas menciona
#       "neutralização" como meta futura.
#   (b) Falso positivo: empresa que *vende* créditos (ex: geradora de energia
#       renovável que emite I-RECs para clientes) — o contexto textual deixa
#       isso claro e o revisor pode corrigir.
#
# Critério de exclusão aplicado em código:
#   - Se a resposta da seção 5.1 contiver "Não" como resposta imediata ao campo
#     "A organização possui projetos de compensação?", a empresa é marcada NÃO
#     a menos que haja evidência forte nas outras seções.

# ── Camada 1: Créditos / mercado externo de carbono ──────────────────────────
PADROES_CREDITO = [

    # ---- Terminologia genérica de crédito de carbono ----
    (r"cr[eé]dito[s]?\s+de\s+carbono",          "crédito de carbono"),
    (r"carbon\s+credit",                          "carbon credit"),
    (r"carbon\s+offset",                          "carbon offset"),
    (r"\boffset[s]?\b",                           "offset"),
    (r"aposentadoria\s+de\s+cr[eé]dito",          "aposentadoria de crédito"),
    (r"retirement\s+of\s+credit",                 "retirement of credit"),
    (r"compra\s+de\s+cr[eé]dito",                 "compra de crédito"),
    (r"aquisi[cç][aã]o\s+de\s+cr[eé]dito",        "aquisição de crédito"),
    (r"cr[eé]dito[s]?\s+de\s+carbono\s+voluntário","crédito voluntário"),

    # ---- Padrões internacionais de certificação ----
    (r"\bverra\b",                                "Verra"),
    (r"\bVCS\b",                                  "VCS (Verra)"),
    (r"verified\s+carbon\s+standard",             "Verified Carbon Standard"),
    (r"\bgold\s+standard\b",                      "Gold Standard"),
    (r"\bCCB\b",                                  "CCB"),
    (r"climate.community.biodiversity",           "CCB Standards"),
    (r"\bREDD\b",                                 "REDD+"),
    (r"\bREDD\+",                                 "REDD+"),
    (r"\bACR\b",                                  "ACR"),
    (r"american\s+carbon\s+registry",             "American Carbon Registry"),
    (r"\bCAR\b",                                  "Climate Action Reserve"),
    (r"climate\s+action\s+reserve",               "Climate Action Reserve"),
    (r"\bCORSIA\b",                               "CORSIA"),
    (r"\bPCF\b",                                  "PCF"),
    (r"\bCBio\b",                                 "CBio"),
    (r"\bRBIO\b",                                 "RBIO"),
    (r"\bRBIO3\b",                                "RBIO3"),

    # ---- MDL / CDM (Mecanismo de Desenvolvimento Limpo) ----
    (r"\bMDL\b",                                  "MDL"),
    (r"mecanismo\s+de\s+desenvolvimento\s+limpo", "MDL"),
    (r"\bCDM\b",                                  "CDM"),
    (r"clean\s+development\s+mechanism",          "CDM"),
    (r"\bCER[s]?\b",                              "CER (MDL)"),
    (r"certified\s+emission\s+reduction",         "CER (MDL)"),

    # ---- VER — Voluntary Emission Reductions ----
    (r"\bVER[s]?\b",                              "VER"),
    (r"voluntary\s+emission\s+reduction",         "VER"),

    # ---- Energia renovável certificada ----
    (r"\bI.?REC[s]?\b",                           "I-REC"),
    (r"international\s+rec\b",                    "I-REC"),
    (r"\bREC[s]?\b(?!\w)",                        "REC"),
    (r"certificado[s]?\s+de\s+energia\s+renov[aá]vel", "Certificado de Energia Renovável"),
    (r"\bGO\b",                                   "Guarantees of Origin (GO)"),
    (r"guarantee[s]?\s+of\s+origin",              "Guarantees of Origin"),

    # ---- Programas/plataformas de crédito brasileiros ----
    (r"\bSBCE\b",                                 "SBCE"),
    (r"sistema\s+brasileiro\s+de\s+com[eé]rcio\s+de\s+emiss[oõ]es","SBCE"),
    (r"mercado\s+de\s+carbono\s+regulado",        "Mercado regulado"),
    (r"mercado\s+regulado",                       "Mercado regulado"),
    (r"bolsa\s+de\s+carbono",                     "Bolsa de carbono"),
    (r"b3\s+carbono",                             "B3 Carbono"),

    # ---- Termos de compra/aquisição no mercado voluntário ----
    (r"mercado\s+volunt[aá]rio\s+de\s+carbono",  "Mercado voluntário de carbono"),
    (r"compensa[cç][aã]o\s+volunt[aá]ria",       "Compensação voluntária"),
    (r"neutraliza[cç][aã]o\s+(?:via|por|atrav[eé]s\s+de)\s+cr[eé]dito", "Neutralização via crédito"),
    (r"cr[eé]dito[s]?\s+(?:de\s+)?florest",      "Crédito florestal"),
    (r"cr[eé]dito[s]?\s+(?:de\s+)?REDD",         "Crédito REDD"),
]

# ── Camada 2: Indicadores contextuais (só validam se Camada 1 > 0) ───────────
PADROES_CONTEXTUAIS = [
    (r"neutraliza[cç][aã]o\s+(?:de\s+)?(?:emiss[oõ]es|carbono)", "neutralização de emissões"),
    (r"carbono\s+neutro",                         "carbono neutro"),
    (r"carbon\s+neutral",                         "carbon neutral"),
    (r"net.?zero",                                "net zero"),
    (r"compensa[cç][aã]o\s+de\s+emiss[oõ]es",    "compensação de emissões"),
    (r"projetos?\s+de\s+compensa[cç][aã]o",       "projeto de compensação"),
]

# ── Padrões de negação — indicam que a empresa NÃO compra créditos ───────────
# Aplicados apenas à resposta direta da seção 5.1 (primeiro parágrafo)
PADROES_NEGACAO_5_1 = [
    r"n[aã]o\s+(?:possui|realiza|aplica|utiliza|tem|h[aá])\s+projeto",
    r"n[aã]o\s+(?:compra|adquire|utiliza)\s+cr[eé]dito",
    r"nenhum\s+projeto\s+de\s+compensa[cç][aã]o",
    r"sem\s+projetos?\s+de\s+compensa[cç][aã]o",
    r"a\s+organiza[cç][aã]o\s+n[aã]o\s+possui\s+projetos?\s+de\s+compensa[cç][aã]o",
]


# ── Extração de texto por página ──────────────────────────────────────────────

def extrair_paginas(caminho_pdf: Path) -> list:
    """Retorna lista de strings, uma por página."""
    try:
        doc = fitz.open(str(caminho_pdf))
        textos = [p.get_text() for p in doc.pages()]
        doc.close()
        return textos
    except Exception as e:
        return [f"__ERRO__: {e}"]


def contar_paginas(caminho_pdf: Path) -> int:
    try:
        doc = fitz.open(str(caminho_pdf))
        n = len(doc)
        doc.close()
        return n
    except Exception:
        return 0


# ── Extração de Nome Fantasia ─────────────────────────────────────────────────

def extrair_nome_fantasia(paginas: list) -> str:
    texto_busca = "\n".join(paginas[:3]) if len(paginas) >= 3 else "\n".join(paginas)
    padroes = [
        r"[Nn]ome\s+[Ff]antasia\s*[:\n]\s*([^\n]{2,100})",
        r"[Nn]ome\s+[Ff]antasia\s{2,}([^\n]{2,100})",
    ]
    for p in padroes:
        m = re.search(p, texto_busca)
        if m:
            val = m.group(1).strip()
            # Ignora linhas que são rótulos de outros campos
            if val and not re.match(
                r"(?:CNPJ|CPF|Setor|Contato|E-mail|Endere[cç]o|Telefone|CEP|Cidade|UF|Pa[ií]s)",
                val, re.IGNORECASE
            ):
                return sanitizar(val)
    return ""


# ── Extração das seções 4.3, 5.1, 5.2 ────────────────────────────────────────

# Mapeamento de seção para padrão regex de início
SECOES_ALVO = {
    "4.3": re.compile(
        r"4[\.\s]+3[\s\.\-–—:]+(?:descri[cç][aã]o\s+de\s+estrat[eé]gias|estrat[eé]gias\s+e\s+projetos)",
        re.IGNORECASE
    ),
    "5.1": re.compile(
        r"5[\.\s]+1[\s\.\-–—:]+(?:compensa[cç][aã]o\s+de\s+emiss[oõ]es|a\s+organiza[cç][aã]o\s+possui\s+projetos\s+de\s+compensa[cç][aã]o)",
        re.IGNORECASE
    ),
    "5.2": re.compile(
        r"5[\.\s]+2[\s\.\-–—:]+(?:redu[cç][oõ]es\s+de\s+emiss[oõ]es|a\s+organiza[cç][aã]o\s+possui\s+projetos\s+de\s+redu[cç][aã]o)",
        re.IGNORECASE
    ),
}

# Padrão que marca o início da próxima seção (para delimitar o fim da atual)
PROXIMA_SECAO = re.compile(
    r"\n\s*(?:\d+[\.\s]+\d+[\.\s]+\d+|\d+[\.\s]+\d+)\s*[\.\-–—:]?\s+[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇ]",
    re.MULTILINE
)


def extrair_secoes(paginas: list) -> dict:
    """
    Extrai o conteúdo textual das seções 4.3, 5.1 e 5.2.
    Estratégia:
      1. Concatena todo o texto do PDF.
      2. Remove bloco 'Informações institucionais' (página 2) para evitar
         falsos positivos com o nome da empresa.
      3. Localiza cada seção pelo regex de início e captura até o início
         da próxima seção numerada.
    Fallback: se nenhuma seção for encontrada, retorna texto completo.
    """
    texto_completo = "\n".join(paginas)

    # Remove bloco de informações institucionais
    texto_completo = re.sub(
        r"Informa[cç][oõ]es\s+institucionais.*?(?=\n\s*\d+[\.\s]+\d+)",
        "",
        texto_completo,
        flags=re.IGNORECASE | re.DOTALL
    )

    secoes_encontradas = {}
    for nome_secao, padrao_inicio in SECOES_ALVO.items():
        m = padrao_inicio.search(texto_completo)
        if not m:
            continue
        inicio = m.start()
        # Busca próxima seção numerada a partir do fim do match
        resto = texto_completo[m.end():]
        m_fim = PROXIMA_SECAO.search(resto)
        if m_fim:
            secao_texto = texto_completo[inicio: m.end() + m_fim.start()]
        else:
            # Fallback: pega até 3000 caracteres
            secao_texto = texto_completo[inicio: inicio + 3000]
        secoes_encontradas[nome_secao] = secao_texto.strip()

    return secoes_encontradas, texto_completo


# ── Detecção de créditos de carbono ──────────────────────────────────────────

def extrair_contexto(texto: str, pos: int) -> str:
    """Retorna trecho de TAMANHO_CONTEXTO chars ao redor da posição."""
    inicio = max(0, pos - TAMANHO_CONTEXTO // 2)
    fim = min(len(texto), pos + TAMANHO_CONTEXTO // 2)
    trecho = texto[inicio:fim].replace("\n", " ")
    return f"...{trecho}..."


def verificar_negacao_5_1(texto_5_1: str) -> bool:
    """Retorna True se a seção 5.1 nega explicitamente ter projetos de compensação."""
    # Analisa apenas os primeiros 400 caracteres (resposta direta ao campo)
    trecho = texto_5_1[:400].lower()
    for p in PADROES_NEGACAO_5_1:
        if re.search(p, trecho, re.IGNORECASE):
            return True
    return False


def analisar_creditos(secoes: dict, texto_completo: str) -> dict:
    """
    Analisa as seções extraídas e retorna:
      - compra_creditos: "SIM" | "NÃO" | "POSSÍVEL" | "ERRO"
      - padroes_encontrados: lista de rótulos dos padrões detectados
      - secoes_com_evidencia: lista de seções onde foram encontrados
      - contextos: lista de trechos contextuais (máx. 4)
      - negacao_5_1: bool — 5.1 nega explicitamente
    """
    resultado = {
        "compra_creditos":       "NÃO",
        "padroes_encontrados":   [],
        "secoes_com_evidencia":  [],
        "contextos":             [],
        "negacao_5_1":           False,
        "secoes_analisadas":     [],
    }

    if not secoes:
        # Sem seções estruturadas → usa texto completo como fallback
        texto_analise = {"(texto completo)": texto_completo}
    else:
        texto_analise = secoes
        resultado["secoes_analisadas"] = list(secoes.keys())

    # Verifica negação explícita em 5.1
    if "5.1" in secoes:
        resultado["negacao_5_1"] = verificar_negacao_5_1(secoes["5.1"])

    achados_camada1 = []
    achados_camada2 = []

    for nome_secao, texto in texto_analise.items():
        texto_lower = texto.lower()

        # Camada 1
        for padrao, rotulo in PADROES_CREDITO:
            for m in re.finditer(padrao, texto_lower):
                contexto = extrair_contexto(texto, m.start())
                achados_camada1.append({
                    "rotulo": rotulo,
                    "secao": nome_secao,
                    "contexto": sanitizar(contexto),
                })

        # Camada 2
        for padrao, rotulo in PADROES_CONTEXTUAIS:
            for m in re.finditer(padrao, texto_lower):
                contexto = extrair_contexto(texto, m.start())
                achados_camada2.append({
                    "rotulo": rotulo,
                    "secao": nome_secao,
                    "contexto": sanitizar(contexto),
                })

    # ── Decisão de classificação ──────────────────────────────────────────
    padroes_unicos_c1 = list(dict.fromkeys(a["rotulo"] for a in achados_camada1))
    padroes_unicos_c2 = list(dict.fromkeys(a["rotulo"] for a in achados_camada2))
    secoes_c1 = list(dict.fromkeys(a["secao"] for a in achados_camada1))
    secoes_c2 = list(dict.fromkeys(a["secao"] for a in achados_camada2))

    # Constrói lista de contextos únicos (máx. 4)
    todos_contextos = [a["contexto"] for a in achados_camada1 + achados_camada2]
    contextos_unicos = list(dict.fromkeys(todos_contextos))[:4]

    resultado["contextos"] = contextos_unicos

    if padroes_unicos_c1:
        # Tem evidência direta de créditos externos
        if resultado["negacao_5_1"] and len(padroes_unicos_c1) == 1 and padroes_unicos_c1[0] in ("offset", "REC"):
            # Apenas termos ambíguos + negação em 5.1 → classificar como POSSÍVEL
            resultado["compra_creditos"] = "POSSÍVEL"
        else:
            resultado["compra_creditos"] = "SIM"
        resultado["padroes_encontrados"] = padroes_unicos_c1 + [
            f"[contexto: {r}]" for r in padroes_unicos_c2
        ]
        resultado["secoes_com_evidencia"] = list(dict.fromkeys(secoes_c1 + secoes_c2))

    elif padroes_unicos_c2 and not resultado["negacao_5_1"]:
        # Só termos contextuais, sem negação → POSSÍVEL (requer revisão)
        resultado["compra_creditos"] = "POSSÍVEL"
        resultado["padroes_encontrados"] = [f"[contexto: {r}]" for r in padroes_unicos_c2]
        resultado["secoes_com_evidencia"] = secoes_c2

    # Se negação em 5.1 e sem Camada 1 forte → NÃO (já é o default)

    return resultado


# ── Análise de todos os PDFs ──────────────────────────────────────────────────

def analisar_pasta(pasta: str) -> list:
    pasta_path = Path(pasta)
    pdfs = sorted(pasta_path.glob("**/*.pdf"))
    if not pdfs:
        print(f"⚠️  Nenhum PDF encontrado em: {pasta}")
        return []

    print(f"📂 {len(pdfs)} PDFs encontrados em {pasta}")
    resultados = []

    iterador = tqdm(pdfs, desc="Analisando PDFs") if TQDM_DISPONIVEL else pdfs

    for pdf_path in iterador:
        if not TQDM_DISPONIVEL:
            print(f"  → {pdf_path.name}")

        paginas = extrair_paginas(pdf_path)

        if paginas and paginas[0].startswith("__ERRO__"):
            resultados.append({
                "arquivo": sanitizar(pdf_path.name),
                "empresa": sanitizar(pdf_path.stem),
                "compra_creditos": "ERRO",
                "padroes_encontrados": "",
                "secoes_com_evidencia": "",
                "secoes_analisadas": "",
                "negacao_5_1": "",
                "contexto_1": paginas[0],
                "contexto_2": "", "contexto_3": "", "contexto_4": "",
                "paginas_analisadas": 0,
            })
            continue

        nome_empresa = extrair_nome_fantasia(paginas) or sanitizar(pdf_path.stem)
        secoes, texto_completo = extrair_secoes(paginas)
        analise = analisar_creditos(secoes, texto_completo)

        resultados.append({
            "arquivo":              sanitizar(pdf_path.name),
            "empresa":              nome_empresa,
            "compra_creditos":      analise["compra_creditos"],
            "padroes_encontrados":  sanitizar(" | ".join(analise["padroes_encontrados"])),
            "secoes_com_evidencia": sanitizar(" | ".join(analise["secoes_com_evidencia"])),
            "secoes_analisadas":    sanitizar(" | ".join(analise["secoes_analisadas"])) or "fallback (texto completo)",
            "negacao_5_1":          "SIM" if analise["negacao_5_1"] else "NÃO",
            "contexto_1":           analise["contextos"][0] if len(analise["contextos"]) > 0 else "",
            "contexto_2":           analise["contextos"][1] if len(analise["contextos"]) > 1 else "",
            "contexto_3":           analise["contextos"][2] if len(analise["contextos"]) > 2 else "",
            "contexto_4":           analise["contextos"][3] if len(analise["contextos"]) > 3 else "",
            "paginas_analisadas":   contar_paginas(pdf_path),
        })

    return resultados


# ── Geração do Excel ──────────────────────────────────────────────────────────

def gerar_excel(resultados: list, caminho_saida: str):
    wb = openpyxl.Workbook()

    COR_CAB      = "1F4E79"
    COR_SIM      = "C6EFCE"
    COR_NAO      = "F4CCCC"
    COR_POSSIVEL = "FFF2CC"
    COR_ERRO     = "FCE4D6"
    COR_ALT      = "EBF3FB"

    f_cab  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    f_sim  = Font(name="Arial", bold=True, color="375623")
    f_nao  = Font(name="Arial", color="7F0000")
    f_pos  = Font(name="Arial", bold=True, color="7F6000")
    f_norm = Font(name="Arial", size=10)
    al_cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_esq = Alignment(horizontal="left", vertical="top", wrap_text=True)
    borda  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── Aba Resultados ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resultados"

    cabecalhos = [
        "Arquivo PDF",
        "Nome Fantasia (Empresa)",
        "Compra Créditos de Carbono?",
        "Revisado?",
        "Padrões / Tipos Identificados",
        "Seções com Evidência",
        "Seções Analisadas",
        "5.1 Negação Explícita?",
        "Páginas",
        "Contexto 1",
        "Contexto 2",
        "Contexto 3",
        "Contexto 4",
    ]
    chaves = [
        "arquivo", "empresa", "compra_creditos", None,
        "padroes_encontrados", "secoes_com_evidencia", "secoes_analisadas",
        "negacao_5_1", "paginas_analisadas",
        "contexto_1", "contexto_2", "contexto_3", "contexto_4",
    ]

    for ci, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=ci, value=cab)
        c.font = f_cab
        c.fill = PatternFill("solid", start_color=COR_CAB)
        c.alignment = al_cen
        c.border = borda

    for ri, res in enumerate(resultados, 2):
        cor_linha = COR_ALT if ri % 2 == 0 else "FFFFFF"
        for ci, chave in enumerate(chaves, 1):
            valor = res.get(chave, "") if chave else ""
            if isinstance(valor, str):
                valor = sanitizar(valor)
            c = ws.cell(row=ri, column=ci, value=valor)
            c.font = f_norm
            c.border = borda
            c.fill = PatternFill("solid", start_color=cor_linha)

            if chave == "compra_creditos":
                if valor == "SIM":
                    c.fill = PatternFill("solid", start_color=COR_SIM)
                    c.font = f_sim
                elif valor == "NÃO":
                    c.fill = PatternFill("solid", start_color=COR_NAO)
                    c.font = f_nao
                elif valor == "POSSÍVEL":
                    c.fill = PatternFill("solid", start_color=COR_POSSIVEL)
                    c.font = f_pos
                elif valor == "ERRO":
                    c.fill = PatternFill("solid", start_color=COR_ERRO)
                c.alignment = al_cen
            elif chave in ("paginas_analisadas", "negacao_5_1"):
                c.alignment = al_cen
            else:
                c.alignment = al_esq

    larguras = [25, 45, 24, 12, 55, 22, 22, 18, 9, 90, 90, 90, 90]
    for i, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    for row in ws.iter_rows(min_row=2, max_row=len(resultados) + 1):
        ws.row_dimensions[row[0].row].height = 60

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Aba Resumo ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")

    total     = len(resultados)
    com_sim   = sum(1 for r in resultados if r["compra_creditos"] == "SIM")
    possivel  = sum(1 for r in resultados if r["compra_creditos"] == "POSSÍVEL")
    com_nao   = sum(1 for r in resultados if r["compra_creditos"] == "NÃO")
    com_erro  = sum(1 for r in resultados if r["compra_creditos"] == "ERRO")

    ws2["A1"] = "RESUMO — COMPENSAÇÃO POR CRÉDITOS DE CARBONO (RPE 2024)"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")

    dados_resumo = [
        ("Total de PDFs analisados:",                    total),
        ("Compram créditos de carbono (SIM):",           com_sim),
        ("% que compram créditos:",                      f"=B4/B3" if total > 0 else "N/A"),
        ("Requerem revisão (POSSÍVEL):",                 possivel),
        ("Não compram créditos (NÃO):",                  com_nao),
        ("Erros de leitura:",                            com_erro),
        ("Data da análise:",                             datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for i, (label, valor) in enumerate(dados_resumo, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
        c = ws2.cell(row=i, column=2, value=valor)
        c.font = Font(name="Arial", size=10)
        if i == 5 and total > 0:
            c.number_format = "0.0%"

    # Tabela de padrões mais frequentes
    ws2["A12"] = "PADRÕES / TIPOS DE CRÉDITO MAIS ENCONTRADOS"
    ws2["A12"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")

    contagem_padroes: dict = {}
    for r in resultados:
        for p in r["padroes_encontrados"].split(" | "):
            p = p.strip().lstrip("[contexto: ").rstrip("]")
            if p:
                contagem_padroes[p] = contagem_padroes.get(p, 0) + 1
    padroes_ord = sorted(contagem_padroes.items(), key=lambda x: x[1], reverse=True)

    for col, title in [("A", "Padrão / Tipo de Crédito"), ("B", "Nº de Empresas")]:
        c = ws2[f"{col}13"]
        c.value = title
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COR_CAB)
        c.alignment = al_cen

    for i, (sw, cnt) in enumerate(padroes_ord, 14):
        cor = COR_ALT if i % 2 == 0 else "FFFFFF"
        ws2[f"A{i}"] = sanitizar(sw)
        ws2[f"B{i}"] = cnt
        for col in ("A", "B"):
            ws2[f"{col}{i}"].fill = PatternFill("solid", start_color=cor)
            ws2[f"{col}{i}"].font = Font(name="Arial", size=10)
        ws2[f"B{i}"].alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 20

    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    print(f"\n✅ Planilha salva em: {caminho_saida}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Analisa PDFs de inventários GEE — Compensação por créditos de carbono."
    )
    parser.add_argument("--pasta",  "-p", default=PASTA_PDFS,
                        help="Pasta com os PDFs (default: ./pdfs)")
    parser.add_argument("--saida",  "-s", default=None,
                        help="Caminho do Excel de saída")
    args = parser.parse_args()

    if args.saida is None:
        data_hoje = datetime.now().strftime("%Y-%m-%d")
        args.saida = str(Path(args.pasta) / f"creditos_carbono_{data_hoje}.xlsx")

    print("=" * 65)
    print("  ANALISADOR DE INVENTÁRIOS GEE — CRÉDITOS DE CARBONO")
    print("=" * 65)

    resultados = analisar_pasta(args.pasta)
    if not resultados:
        print("Nenhum resultado para salvar.")
        return

    total    = len(resultados)
    sim      = sum(1 for r in resultados if r["compra_creditos"] == "SIM")
    possivel = sum(1 for r in resultados if r["compra_creditos"] == "POSSÍVEL")
    erros    = sum(1 for r in resultados if r["compra_creditos"] == "ERRO")

    print(f"\n📊 Resumo:")
    print(f"   Total de PDFs:          {total}")
    print(f"   Compram créditos (SIM): {sim} ({100*sim/total:.1f}%)")
    print(f"   Requerem revisão:       {possivel} ({100*possivel/total:.1f}%)")
    print(f"   Não compram (NÃO):      {total - sim - possivel - erros}")
    print(f"   Erros de leitura:       {erros}")

    gerar_excel(resultados, args.saida)


if __name__ == "__main__":
    main()
