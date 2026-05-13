"""
Extrator de CNPJ — Inventários de GEE (RPE 2024)
=================================================
Lê apenas a segunda página de cada PDF e extrai o CNPJ da empresa.

Estratégia de extração:
    1. Localiza o rótulo "CNPJ" na página 2 e captura o número logo após.
    2. Fallback: varre toda a página pelo padrão numérico XX.XXX.XXX/XXXX-XX.
    3. Fallback 2: busca 14 dígitos consecutivos sem formatação.
"""

import os
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime

try:
    import fitz  # PyMuPDF
except ImportError:
    print("❌ PyMuPDF não encontrado. Instale com: pip install pymupdf")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
    TQDM_DISPONIVEL = True
except ImportError:
    TQDM_DISPONIVEL = False

# ── Configuração ──────────────────────────────────────────────────────────────
PASTA_PDFS = os.environ.get("PASTA_PDFS", "./pdfs")

_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\uFFFE\uFFFF]")

def sanitizar(texto: str) -> str:
    if not isinstance(texto, str):
        return str(texto) if texto is not None else ""
    limpo = _ILLEGAL_CHARS_RE.sub(" ", texto)
    return re.sub(r" {2,}", " ", limpo).strip()


# ── Padrões de CNPJ ───────────────────────────────────────────────────────────

# Padrão com máscara: 12.345.678/0001-90
CNPJ_FORMATADO = re.compile(
    r"\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\s/]?\d{4}[-\s]?\d{2}"
)

# Padrão sem máscara: 12345678000190
CNPJ_SEM_MASCARA = re.compile(r"\b\d{14}\b")


def formatar_cnpj(digitos: str) -> str:
    """Aplica a máscara padrão XX.XXX.XXX/XXXX-XX."""
    d = re.sub(r"\D", "", digitos)
    if len(d) != 14:
        return digitos
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def validar_cnpj(digitos: str) -> bool:
    """
    Valida os dígitos verificadores do CNPJ.
    Retorna True se o CNPJ for matematicamente válido.
    """
    d = re.sub(r"\D", "", digitos)
    if len(d) != 14:
        return False
    if len(set(d)) == 1:
        return False  # Sequências como 00000000000000 são inválidas

    def calcular_digito(d, pesos):
        soma = sum(int(d[i]) * pesos[i] for i in range(len(pesos)))
        resto = soma % 11
        return 0 if resto < 2 else 11 - resto

    pesos1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    pesos2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]

    d1 = calcular_digito(d, pesos1)
    d2 = calcular_digito(d, pesos2)

    return int(d[12]) == d1 and int(d[13]) == d2


# ── Extração de texto da página 2 ────────────────────────────────────────────

def extrair_pagina_2(caminho_pdf: Path) -> str:
    """Retorna o texto da segunda página (índice 1) do PDF."""
    try:
        doc = fitz.open(str(caminho_pdf))
        if len(doc) < 2:
            # PDF tem menos de 2 páginas — tenta a primeira
            texto = doc[0].get_text()
            doc.close()
            return texto
        texto = doc[1].get_text()
        doc.close()
        return texto
    except Exception as e:
        return f"__ERRO__: {e}"


def extrair_nome_fantasia(caminho_pdf: Path) -> str:
    """Extrai o Nome Fantasia das primeiras páginas do PDF."""
    try:
        doc = fitz.open(str(caminho_pdf))
        paginas = [doc[i].get_text() for i in range(min(3, len(doc)))]
        doc.close()
    except Exception:
        return ""

    texto = "\n".join(paginas)
    padroes = [
        r"[Nn]ome\s+[Ff]antasia\s*[:\n]\s*([^\n]{2,100})",
        r"[Nn]ome\s+[Ff]antasia\s{2,}([^\n]{2,100})",
    ]
    for p in padroes:
        m = re.search(p, texto)
        if m:
            val = m.group(1).strip()
            if val and not re.match(
                r"(?:CNPJ|CPF|Setor|Contato|E-mail|Endere[cç]o|Telefone|CEP|Cidade|UF|Pa[ií]s)",
                val, re.IGNORECASE
            ):
                return sanitizar(val)
    return ""


# ── Extração do CNPJ ──────────────────────────────────────────────────────────

def extrair_cnpj(texto_pag2: str) -> dict:
    """
    Tenta extrair o CNPJ do texto da página 2.

    Estratégias (em ordem de prioridade):
        1. Busca o rótulo "CNPJ" e captura o número nos próximos ~60 caracteres.
        2. Varre toda a página pelo padrão formatado XX.XXX.XXX/XXXX-XX.
        3. Varre toda a página por 14 dígitos consecutivos.

    Retorna dict com:
        - cnpj_formatado: string com máscara ou vazia
        - cnpj_digitos:   apenas os 14 dígitos ou vazio
        - valido:         bool — CNPJ passou na validação dos dígitos verificadores
        - metodo:         como foi encontrado
        - observacao:     mensagem de erro/aviso se não encontrado
    """
    resultado = {
        "cnpj_formatado": "",
        "cnpj_digitos":   "",
        "valido":         False,
        "metodo":         "",
        "observacao":     "",
    }

    if texto_pag2.startswith("__ERRO__"):
        resultado["observacao"] = texto_pag2
        return resultado

    # ── Estratégia 1: rótulo CNPJ + número logo após ──────────────────────
    # Cobre formatos como:
    #   "CNPJ\n12.345.678/0001-90"
    #   "CNPJ: 12.345.678/0001-90"
    #   "CNPJ  12345678000190"
    padrao_rotulo = re.compile(
        r"CNPJ\s*[:\-–]?\s*"           # rótulo
        r"(\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\s/]?\d{4}[-\s]?\d{2}|\d{14})",
        re.IGNORECASE
    )
    m = padrao_rotulo.search(texto_pag2)
    if m:
        candidato = m.group(1)
        digitos = re.sub(r"\D", "", candidato)
        if len(digitos) == 14:
            resultado["cnpj_formatado"] = formatar_cnpj(digitos)
            resultado["cnpj_digitos"]   = digitos
            resultado["valido"]         = validar_cnpj(digitos)
            resultado["metodo"]         = "rótulo CNPJ"
            if not resultado["valido"]:
                resultado["observacao"] = "CNPJ encontrado mas inválido (dígitos verificadores)"
            return resultado

    # ── Estratégia 2: padrão formatado em toda a página ───────────────────
    candidatos = CNPJ_FORMATADO.findall(texto_pag2)
    for candidato in candidatos:
        digitos = re.sub(r"\D", "", candidato)
        if len(digitos) == 14:
            resultado["cnpj_formatado"] = formatar_cnpj(digitos)
            resultado["cnpj_digitos"]   = digitos
            resultado["valido"]         = validar_cnpj(digitos)
            resultado["metodo"]         = "padrão formatado (página 2)"
            if not resultado["valido"]:
                resultado["observacao"] = "CNPJ encontrado mas inválido (dígitos verificadores)"
            return resultado

    # ── Estratégia 3: 14 dígitos consecutivos ────────────────────────────
    candidatos_raw = CNPJ_SEM_MASCARA.findall(texto_pag2)
    for candidato in candidatos_raw:
        if validar_cnpj(candidato):
            resultado["cnpj_formatado"] = formatar_cnpj(candidato)
            resultado["cnpj_digitos"]   = candidato
            resultado["valido"]         = True
            resultado["metodo"]         = "14 dígitos consecutivos (fallback)"
            return resultado

    # Não encontrado
    resultado["observacao"] = "CNPJ não encontrado na página 2"
    return resultado


# ── Análise de todos os PDFs ──────────────────────────────────────────────────

def analisar_pasta(pasta: str) -> list:
    pasta_path = Path(pasta)
    pdfs = sorted(pasta_path.glob("**/*.pdf"))
    if not pdfs:
        print(f"⚠️  Nenhum PDF encontrado em: {pasta}")
        return []

    print(f"📂 {len(pdfs)} PDFs encontrados em {pasta}")
    resultados = []

    iterador = tqdm(pdfs, desc="Extraindo CNPJs") if TQDM_DISPONIVEL else pdfs

    for pdf_path in iterador:
        if not TQDM_DISPONIVEL:
            print(f"  → {pdf_path.name}")

        texto_pag2  = extrair_pagina_2(pdf_path)
        nome        = extrair_nome_fantasia(pdf_path) or sanitizar(pdf_path.stem)
        cnpj_info   = extrair_cnpj(texto_pag2)

        resultados.append({
            "arquivo":        sanitizar(pdf_path.name),
            "empresa":        nome,
            "cnpj_formatado": cnpj_info["cnpj_formatado"],
            "cnpj_digitos":   cnpj_info["cnpj_digitos"],
            "valido":         "SIM" if cnpj_info["valido"] else ("NÃO" if cnpj_info["cnpj_digitos"] else ""),
            "metodo":         cnpj_info["metodo"],
            "observacao":     cnpj_info["observacao"],
        })

    return resultados


# ── Geração do Excel ──────────────────────────────────────────────────────────

def gerar_excel(resultados: list, caminho_saida: str):
    wb = openpyxl.Workbook()

    COR_CAB  = "1F4E79"
    COR_SIM  = "C6EFCE"
    COR_NAO  = "F4CCCC"
    COR_ALT  = "EBF3FB"
    COR_ERRO = "FCE4D6"

    f_cab  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    f_sim  = Font(name="Arial", bold=True, color="375623")
    f_nao  = Font(name="Arial", color="7F0000")
    f_norm = Font(name="Arial", size=10)
    al_cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_esq = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    borda  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── Aba Resultados ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resultados"

    cabecalhos = [
        "Arquivo PDF",
        "Nome Fantasia (Empresa)",
        "CNPJ (formatado)",
        "CNPJ (somente dígitos)",
        "CNPJ Válido?",
        "Método de Extração",
        "Observação",
    ]
    chaves = [
        "arquivo", "empresa", "cnpj_formatado",
        "cnpj_digitos", "valido", "metodo", "observacao",
    ]

    for ci, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=ci, value=cab)
        c.font = f_cab
        c.fill = PatternFill("solid", start_color=COR_CAB)
        c.alignment = al_cen
        c.border = borda

    for ri, res in enumerate(resultados, 2):
        cor_linha = COR_ALT if ri % 2 == 0 else "FFFFFF"
        for ci, chave in enumerate(chaves, 1):
            valor = sanitizar(res.get(chave, ""))
            c = ws.cell(row=ri, column=ci, value=valor)
            c.font = f_norm
            c.border = borda

            if chave == "valido":
                if valor == "SIM":
                    c.fill = PatternFill("solid", start_color=COR_SIM)
                    c.font = f_sim
                elif valor == "NÃO":
                    c.fill = PatternFill("solid", start_color=COR_NAO)
                    c.font = f_nao
                elif not valor:
                    c.fill = PatternFill("solid", start_color=COR_ERRO)
                else:
                    c.fill = PatternFill("solid", start_color=cor_linha)
                c.alignment = al_cen
            elif chave in ("cnpj_formatado", "cnpj_digitos", "metodo"):
                c.fill = PatternFill("solid", start_color=cor_linha)
                c.alignment = al_cen
            else:
                c.fill = PatternFill("solid", start_color=cor_linha)
                c.alignment = al_esq

    larguras = [30, 50, 22, 22, 15, 35, 45]
    for i, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2, max_row=len(resultados) + 1):
        ws.row_dimensions[row[0].row].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Aba Resumo ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")

    total         = len(resultados)
    com_cnpj      = sum(1 for r in resultados if r["cnpj_formatado"])
    cnpj_valido   = sum(1 for r in resultados if r["valido"] == "SIM")
    cnpj_invalido = sum(1 for r in resultados if r["valido"] == "NÃO")
    sem_cnpj      = sum(1 for r in resultados if not r["cnpj_formatado"])

    ws2["A1"] = "RESUMO — EXTRAÇÃO DE CNPJ (RPE 2024)"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")

    dados = [
        ("Total de PDFs analisados:",         total),
        ("CNPJs extraídos:",                  com_cnpj),
        ("CNPJs matematicamente válidos:",     cnpj_valido),
        ("CNPJs com dígitos inválidos:",       cnpj_invalido),
        ("PDFs sem CNPJ encontrado:",          sem_cnpj),
        ("Data da análise:",                   datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for i, (label, valor) in enumerate(dados, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=valor).font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 15

    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    print(f"\n✅ Planilha salva em: {caminho_saida}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extrai CNPJ da página 2 dos PDFs de inventários GEE."
    )
    parser.add_argument("--pasta", "-p", default=PASTA_PDFS,
                        help="Pasta com os PDFs (default: ./pdfs)")
    parser.add_argument("--saida", "-s", default=None,
                        help="Caminho do Excel de saída")
    args = parser.parse_args()

    if args.saida is None:
        data_hoje = datetime.now().strftime("%Y-%m-%d")
        args.saida = str(Path(args.pasta) / f"cnpj_empresas_{data_hoje}.xlsx")

    print("=" * 55)
    print("  EXTRATOR DE CNPJ — INVENTÁRIOS GEE (RPE 2024)")
    print("=" * 55)

    resultados = analisar_pasta(args.pasta)
    if not resultados:
        print("Nenhum resultado para salvar.")
        return

    total       = len(resultados)
    com_cnpj    = sum(1 for r in resultados if r["cnpj_formatado"])
    sem_cnpj    = total - com_cnpj

    print(f"\n📊 Resumo:")
    print(f"   Total de PDFs:        {total}")
    print(f"   CNPJs encontrados:    {com_cnpj} ({100*com_cnpj/total:.1f}%)")
    print(f"   Sem CNPJ:             {sem_cnpj}")

    gerar_excel(resultados, args.saida)


if __name__ == "__main__":
    main()
