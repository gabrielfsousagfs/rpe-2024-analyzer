"""
Analisador de Inventários de GEE - Identificação de Softwares/Plataformas
=========================================================================
v4 - Mudanças:
    - Análise restrita às seções 3.1, 3.2, 4.5 e 4.6 do formulário GHG Protocol
    - Seção "Informações institucionais" explicitamente ignorada
    - Removidos: gvces, fgv ces, categoria "Planilha"
    - Mantém extração de Nome Fantasia, e-mail e sanitização
"""

import os
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime

try:
    import fitz  # PyMuPDF
except ImportError:
    print("❌ PyMuPDF não encontrado. Instale com: pip install pymupdf")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
    TQDM_DISPONIVEL = True
except ImportError:
    TQDM_DISPONIVEL = False

# ── Configuração ──────────────────────────────────────────────────────────────
PASTA_PDFS       = os.environ.get("PASTA_PDFS", "./pdfs")
MAX_PAGINAS      = None
TAMANHO_CONTEXTO = 300

_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\uFFFE\uFFFF]")

def sanitizar(texto: str) -> str:
    if not isinstance(texto, str):
        return str(texto) if texto is not None else ""
    limpo = _ILLEGAL_CHARS_RE.sub(" ", texto)
    return re.sub(r" {2,}", " ", limpo).strip()


# ── Softwares externos conhecidos ─────────────────────────────────────────────
SOFTWARES_CONHECIDOS = {

    "Plataforma Brasileira GEE": [
        r"way\s*carbon",
        r"waycarbon",
        r"deep\s*esg",
        r"\bclimas\b",
        r"\becosystem\b",   # produto WayCarbon
        r"\bimaflora\b",
        r"\bidesam\b",
        r"\becam\b",
        r"\bqualidata\b",
        r"\bingee\b",
        r"\bcerensa\b",
        r"\bakvo[\s\-]?esg\b",
        r"\bneutrality\b",
        r"\bemisfera\b",
        r"\bvankka\b",
        r"\barca\s+sustentabilidade\b",
        r"\bsigea\b",
        r"esphera[\.\s]?bi",
        r"\btbl\s*manager\b",
        r"sinai\b",
    ],

    "Plataforma/Sistema GEE Internacional": [
        r"carbon\s*analytics",
        r"\bgreenbiz\b",
        r"\bsphera\b",
        r"\bwatershed\b",
        r"\bpersefoni\b",
        r"\bnormative\b",
        r"plan\s*a\b",
        r"net\s*zero\s*cloud",
        r"salesforce\s*net\s*zero",
        r"microsoft\s*cloud\s*for\s*sustainability",
        r"sap\s*sustainability",
        r"\benablon\b",
        r"\bintelex\b",
        r"\bcority\b",
        r"\becoact\b",
        r"\becodesk\b",
        r"\bcarbonfact\b",
        r"\bclimatiq\b",
        r"\bemitwise\b",
        r"\bcarbonsmart\b",
        r"\bcarbonchain\b",
        r"\bworkiva\b",
        r"diligent\s*esg",
        r"\bbriink\b",
        r"\bmeasurabl\b",
        r"\becometrica\b",
        r"co2\s*logic",
        r"\bcredit360\b",
        r"ul\s*solutions\b.*\bsoftware\b",
    ],

    "ACV / LCA": [
        r"\bsimapro\b",
        r"open\s*lca",
        r"\bgabi\b",
        r"\becoinvent\b",
        r"one\s*click\s*lca",
        r"\becochain\b",
    ],

    "ERP / Sistema Corporativo": [
        r"\bsap\b",
        r"\boracle\b",
        r"jd\s*edwards",
        r"\btotvs\b",
        r"senior\s*sistemas",
        r"\bdatasul\b",
        r"\bprotheus\b",
        r"\blinx\b",
    ],

    "Energia / Utilidades": [
        r"energy\s*star\s*portfolio\s*manager",
        r"retscreen",
        r"ret\s*screen",
        r"\benergyplus\b",
        r"homer\s*energy",
    ],

    "BI / Dados": [
        r"power\s*bi",
        r"\btableau\b",
        r"\bqlik\b",
        r"\blooker\b",
        r"\bmetabase\b",
    ],
}

# Padrões genéricos — capturam softwares não listados acima
PADROES_GENERICOS = [
    r"(?:software|plataforma|sistema|ferramenta|aplicativo)\s+(?:[\w]+\s+){0,3}[\"'\u201c\u201d]?([\w\s\-\.]{3,50}?)[\"'\u201c\u201d]?\s*(?:,|\.|;|\n|para\b|foi\b)",
    r"(?:utilizad[oa]|usad[oa]|empregad[oa]|adotad[oa])\s+(?:o|a|o\s+software|a\s+plataforma|o\s+sistema|a\s+ferramenta)\s+[\"'\u201c\u201d]?([\w\s\-\.]{3,50}?)[\"'\u201c\u201d]?\s*[,\.\n]",
    r"calculad[oa]\s+(?:no|na|pelo|pela|através\s+do|através\s+da)\s+(?:software|plataforma|sistema|ferramenta)\s+[\"'\u201c\u201d]?([\w\s\-\.]{3,50}?)[\"'\u201c\u201d]?\s*[,\.\n]",
    r"[\"'\u201c\u201d]([\w\s\-\.]{3,50})[\"'\u201c\u201d]\s*[,]?\s*(?:software|plataforma|sistema|ferramenta)",
]

PALAVRAS_IGNORAR = {
    "o", "a", "os", "as", "um", "uma", "que", "de", "do", "da",
    "para", "com", "em", "por", "este", "esta", "esse", "essa",
    "dados", "informações", "relatório", "cálculos", "sistema",
    "software", "plataforma", "ferramenta", "ghg", "gee", "co2",
    "protocolo", "protocol", "programa", "brasileiro", "método",
    "metodologia", "ferramenta", "cálculo", "escopo", "emissões",
}


# ── Extração de seções específicas ────────────────────────────────────────────

# Títulos das seções de interesse (com variações de formatação dos PDFs)
SECOES_ALVO = [
    r"3[\.\s]*1[\s\.\:►▶●◆\-]*[Mm][ée]todo\s*e\s*/\s*ou\s*ferramentas\s*intersetoriais",
    r"3[\.\s]*2[\s\.\:►▶●◆\-]*[Mm][ée]todo\s*e\s*/\s*ou\s*ferramentas\s*para\s*setores\s*espec[íi]ficos",
    r"4[\.\s]*5[\s\.\:►▶●◆\-]*[Ii]nforma[çc][õo]es\s*sobre\s*incertezas",
    r"4[\.\s]*6[\s\.\:►▶●◆\-]*[Dd]escri[çc][ãa]o\s*sobre\s*a[çc][õo]es\s*internas",
]

# Títulos que marcam o fim de uma seção de interesse (próximas seções do formulário)
SECOES_FIM = [
    r"3[\.\s]*3[\s\.\:►▶●◆\-]*[Ff]atores\s*de\s*emiss[ãa]o",
    r"4[\.\s]*[1-47-9][\s\.\:►▶●◆\-]",
    r"5[\.\s]*[0-9][\s\.\:►▶●◆\-]",
    r"[Ii]nforma[çc][õo]es\s*institucionais\s*[:：]?",
    r"[Ss]umário\s+[Ee]xecutivo",
    r"[Aa]p[êe]ndice",
]

# Seção a ignorar explicitamente (descrição institucional da empresa)
SECAO_IGNORAR = re.compile(
    r"[Ii]nforma[çc][õo]es\s*institucionais\s*[:：]",
    re.IGNORECASE
)

PADRAO_ALVO   = re.compile("|".join(SECOES_ALVO),   re.IGNORECASE)
PADRAO_FIM    = re.compile("|".join(SECOES_FIM),     re.IGNORECASE)


def extrair_secoes_relevantes(texto_completo: str) -> str:
    """
    Extrai apenas o conteúdo das seções 3.1, 3.2, 4.5 e 4.6 do formulário GHG Protocol.
    Ignora a seção 'Informações institucionais'.
    Retorna a concatenação dessas seções.
    """
    trechos = []
    pos = 0
    tamanho = len(texto_completo)

    while pos < tamanho:
        # Encontra próxima seção de interesse
        m_alvo = PADRAO_ALVO.search(texto_completo, pos)
        if not m_alvo:
            break

        inicio_secao = m_alvo.start()

        # Encontra onde essa seção termina (próxima seção relevante ou de fim)
        m_fim = PADRAO_FIM.search(texto_completo, m_alvo.end())
        fim_secao = m_fim.start() if m_fim else min(inicio_secao + 2000, tamanho)

        trecho = texto_completo[inicio_secao:fim_secao].strip()

        # Ignora se o trecho contiver "Informações institucionais"
        if not SECAO_IGNORAR.search(trecho):
            trechos.append(trecho)

        pos = m_alvo.end()

    return "\n\n".join(trechos)


# ── Extração de texto por página ──────────────────────────────────────────────

def extrair_paginas(caminho_pdf: Path, max_paginas=None) -> list:
    """Retorna lista de strings, uma por página."""
    try:
        doc = fitz.open(str(caminho_pdf))
        paginas = list(doc.pages()) if max_paginas is None else list(doc.pages())[:max_paginas]
        textos = [p.get_text() for p in paginas]
        doc.close()
        return textos
    except Exception as e:
        return [f"__ERRO__: {e}"]


def contar_paginas(caminho_pdf: Path, max_paginas=None) -> int:
    try:
        doc = fitz.open(str(caminho_pdf))
        n = len(doc)
        doc.close()
        return n if max_paginas is None else min(n, max_paginas)
    except Exception:
        return 0


# ── Extração de Nome Fantasia ─────────────────────────────────────────────────

def extrair_nome_fantasia(paginas: list) -> str:
    """
    Busca 'Nome Fantasia' nas primeiras 3 páginas.
    NÃO usa o bloco 'Informações institucionais' para evitar capturar
    a descrição da empresa ao invés do nome.
    """
    texto_busca = "\n".join(paginas[:3]) if len(paginas) >= 3 else "\n".join(paginas)

    # Remove o bloco de Informações institucionais antes de buscar o nome fantasia
    texto_busca = re.sub(
        r"[Ii]nforma[çc][õo]es\s*institucionais\s*[:：].*",
        "",
        texto_busca,
        flags=re.DOTALL
    )

    padroes = [
        r"[Nn]ome\s+[Ff]antasia\s*[:\n]\s*([^\n]{2,100})",
        r"[Nn]ome\s+[Ff]antasia\s{2,}([^\n]{2,100})",
    ]

    for padrao in padroes:
        m = re.search(padrao, texto_busca)
        if m:
            nome = m.group(1).strip()
            nome = re.sub(r"\s+\d{1,2}/\d{1,2}/\d{4}.*$", "", nome).strip()
            nome = re.sub(r"\s+\d+\s*$", "", nome).strip()
            if len(nome) > 1:
                return sanitizar(nome)

    return ""


# ── Extração de E-mail do Responsável ────────────────────────────────────────

_EMAIL_RE = re.compile(r"[\w\.\-\+]+@[\w\.\-]+\.[a-zA-Z]{2,}")

def extrair_email_responsavel(paginas: list) -> str:
    """Busca o e-mail na seção 'Dados do inventário', primeiras 4 páginas."""
    texto_busca = "\n".join(paginas[:4]) if len(paginas) >= 4 else "\n".join(paginas)

    padroes_label = [
        r"[Ee]-?mail\s+do\s+[Rr]espons[áa]vel\s*[:\n]\s*([\w\.\-\+]+@[\w\.\-]+\.[a-zA-Z]{2,})",
        r"[Ee]-?mail\s+do\s+[Rr]espons[áa]vel\s*[:\n][^\n@]{0,80}?([\w\.\-\+]+@[\w\.\-]+\.[a-zA-Z]{2,})",
        r"[Ee]-?mail\s*[:\n]\s*([\w\.\-\+]+@[\w\.\-]+\.[a-zA-Z]{2,})",
    ]

    for padrao in padroes_label:
        m = re.search(padrao, texto_busca)
        if m:
            return sanitizar(m.group(1).strip())

    emails = _EMAIL_RE.findall(texto_busca)
    if emails:
        return sanitizar(emails[0])

    return ""


# ── Detecção de softwares ─────────────────────────────────────────────────────

def encontrar_softwares_conhecidos(texto: str) -> list:
    texto_lower = texto.lower()
    achados = []
    for categoria, padroes in SOFTWARES_CONHECIDOS.items():
        for padrao in padroes:
            for match in re.finditer(padrao, texto_lower, re.IGNORECASE):
                inicio = max(0, match.start() - TAMANHO_CONTEXTO)
                fim    = min(len(texto), match.end() + TAMANHO_CONTEXTO)
                contexto = texto[inicio:fim].replace("\n", " ").strip()
                nome     = texto[match.start():match.end()].strip()
                achados.append({"software": nome, "categoria": categoria, "contexto": contexto})
    return achados


def encontrar_softwares_genericos(texto: str) -> list:
    achados = []
    for padrao in PADROES_GENERICOS:
        for match in re.finditer(padrao, texto, re.IGNORECASE):
            try:
                nome = match.group(1).strip()
            except IndexError:
                continue
            if len(nome) < 3 or len(nome) > 60 or nome.lower() in PALAVRAS_IGNORAR:
                continue
            inicio = max(0, match.start() - TAMANHO_CONTEXTO)
            fim    = min(len(texto), match.end() + TAMANHO_CONTEXTO)
            contexto = texto[inicio:fim].replace("\n", " ").strip()
            achados.append({"software": nome, "categoria": "Identificado automaticamente", "contexto": contexto})
    return achados


def consolidar_achados(achados: list) -> dict:
    vistos = set()
    r = {"softwares": [], "categorias": [], "contextos": []}
    for a in achados:
        chave = a["software"].lower().strip()
        if chave not in vistos:
            vistos.add(chave)
            r["softwares"].append(a["software"])
            r["categorias"].append(a["categoria"])
            r["contextos"].append(a["contexto"][:500])
    return r


# ── Análise da pasta ──────────────────────────────────────────────────────────

def analisar_pasta(pasta: str, max_paginas=None) -> list:
    pdfs = sorted(Path(pasta).glob("**/*.pdf"))
    if not pdfs:
        print(f"⚠️  Nenhum PDF encontrado em: {pasta}")
        return []

    print(f"\n📂 Pasta: {pasta}")
    print(f"📄 PDFs encontrados: {len(pdfs)}\n")

    resultados = []
    iterador = tqdm(pdfs, desc="Analisando PDFs", unit="pdf") if TQDM_DISPONIVEL else pdfs

    for pdf_path in iterador:
        paginas = extrair_paginas(pdf_path, max_paginas)

        if paginas and paginas[0].startswith("__ERRO__"):
            resultados.append({
                "arquivo":            sanitizar(pdf_path.name),
                "empresa":            "",
                "email_responsavel":  "",
                "usa_software":       "ERRO",
                "softwares":          "",
                "categorias":         "",
                "secoes_encontradas": "",
                "contexto_1":         sanitizar(paginas[0]),
                "contexto_2":         "",
                "contexto_3":         "",
                "total_softwares":    0,
                "paginas_analisadas": 0,
            })
            continue

        texto_completo = "\n".join(paginas)

        # Metadados (nome fantasia e email buscam no texto completo, não só nas seções)
        nome_empresa      = extrair_nome_fantasia(paginas)
        email_responsavel = extrair_email_responsavel(paginas)

        if not nome_empresa:
            nome_empresa = sanitizar(pdf_path.stem)

        # Extrai apenas seções relevantes para detecção de software
        texto_secoes = extrair_secoes_relevantes(texto_completo)
        secoes_encontradas = "SIM" if texto_secoes.strip() else "NÃO"

        # Se não encontrou seções estruturadas (PDF muito compactado), usa texto completo
        # mas remove o bloco de Informações Institucionais
        if not texto_secoes.strip():
            texto_secoes = re.sub(
                r"[Ii]nforma[çc][õo]es\s*institucionais\s*[:：].*?(?=\n\n|\Z)",
                "",
                texto_completo,
                flags=re.DOTALL
            )

        achados = encontrar_softwares_conhecidos(texto_secoes) + encontrar_softwares_genericos(texto_secoes)
        c = consolidar_achados(achados)

        resultados.append({
            "arquivo":            sanitizar(pdf_path.name),
            "empresa":            nome_empresa,
            "email_responsavel":  email_responsavel,
            "usa_software":       "SIM" if c["softwares"] else "NÃO",
            "softwares":          sanitizar(" | ".join(c["softwares"])),
            "categorias":         sanitizar(" | ".join(c["categorias"])),
            "secoes_encontradas": secoes_encontradas,
            "contexto_1":         sanitizar(c["contextos"][0]) if len(c["contextos"]) > 0 else "",
            "contexto_2":         sanitizar(c["contextos"][1]) if len(c["contextos"]) > 1 else "",
            "contexto_3":         sanitizar(c["contextos"][2]) if len(c["contextos"]) > 2 else "",
            "total_softwares":    len(c["softwares"]),
            "paginas_analisadas": contar_paginas(pdf_path, max_paginas),
        })

    return resultados


# ── Geração da planilha Excel ─────────────────────────────────────────────────

def gerar_excel(resultados: list, caminho_saida: str):
    wb = openpyxl.Workbook()

    COR_CAB  = "1F4E79"
    COR_SIM  = "C6EFCE"
    COR_NAO  = "F4CCCC"
    COR_ERRO = "FCE4D6"
    COR_ALT  = "EBF3FB"

    f_cab  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    f_sim  = Font(name="Arial", bold=True, color="375623")
    f_nao  = Font(name="Arial", color="7F0000")
    f_norm = Font(name="Arial", size=10)
    al_cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_esq = Alignment(horizontal="left", vertical="top", wrap_text=True)
    borda  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── Aba Resultados ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resultados"

    cabecalhos = [
        "Arquivo PDF",
        "Nome Fantasia (Empresa)",
        "E-mail Responsável",
        "Usa Software Externo?",
        "Softwares / Plataformas",
        "Categorias",
        "Seções 3.1/3.2/4.5/4.6 encontradas?",
        "Total",
        "Páginas",
        "Contexto 1",
        "Contexto 2",
        "Contexto 3",
    ]
    chaves = [
        "arquivo",
        "empresa",
        "email_responsavel",
        "usa_software",
        "softwares",
        "categorias",
        "secoes_encontradas",
        "total_softwares",
        "paginas_analisadas",
        "contexto_1",
        "contexto_2",
        "contexto_3",
    ]

    for ci, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=ci, value=cab)
        c.font = f_cab
        c.fill = PatternFill("solid", start_color=COR_CAB)
        c.alignment = al_cen
        c.border = borda

    for ri, res in enumerate(resultados, 2):
        cor_linha = COR_ALT if ri % 2 == 0 else "FFFFFF"
        for ci, chave in enumerate(chaves, 1):
            valor = res.get(chave, "")
            if isinstance(valor, str):
                valor = sanitizar(valor)
            c = ws.cell(row=ri, column=ci, value=valor)
            c.font = f_norm
            c.border = borda
            c.fill = PatternFill("solid", start_color=cor_linha)

            if chave == "usa_software":
                if valor == "SIM":
                    c.fill = PatternFill("solid", start_color=COR_SIM)
                    c.font = f_sim
                elif valor == "NÃO":
                    c.fill = PatternFill("solid", start_color=COR_NAO)
                    c.font = f_nao
                elif valor == "ERRO":
                    c.fill = PatternFill("solid", start_color=COR_ERRO)
                c.alignment = al_cen
            elif chave in ("total_softwares", "paginas_analisadas", "secoes_encontradas"):
                c.alignment = al_cen
            else:
                c.alignment = al_esq

    larguras = [25, 42, 32, 18, 50, 35, 22, 8, 8, 80, 80, 80]
    for i, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    for row in ws.iter_rows(min_row=2, max_row=len(resultados) + 1):
        ws.row_dimensions[row[0].row].height = 55

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Aba Resumo ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    total      = len(resultados)
    com_sw     = sum(1 for r in resultados if r["usa_software"] == "SIM")
    sem_sw     = sum(1 for r in resultados if r["usa_software"] == "NÃO")
    com_erro   = sum(1 for r in resultados if r["usa_software"] == "ERRO")
    com_email  = sum(1 for r in resultados if r.get("email_responsavel", ""))
    com_secoes = sum(1 for r in resultados if r.get("secoes_encontradas") == "SIM")

    contagem_sw = {}
    for r in resultados:
        for sw in r["softwares"].split(" | "):
            sw = sw.strip()
            if sw:
                contagem_sw[sw] = contagem_sw.get(sw, 0) + 1
    sw_ord = sorted(contagem_sw.items(), key=lambda x: x[1], reverse=True)

    ws2["A1"] = "RESUMO DA ANÁLISE — INVENTÁRIOS GEE 2024 (v4 — seções 3.1/3.2/4.5/4.6)"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")

    dados_resumo = [
        ("Total de PDFs analisados:", total),
        ("PDFs com seções estruturadas encontradas:", com_secoes),
        ("Empresas com software externo identificado:", com_sw),
        ("% que usam software externo:", "=B5/B3" if total > 0 else "N/A"),
        ("Empresas sem software externo:", sem_sw),
        ("E-mails de responsáveis encontrados:", com_email),
        ("PDFs com erro de leitura:", com_erro),
        ("Data da análise:", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for i, (label, valor) in enumerate(dados_resumo, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
        c = ws2.cell(row=i, column=2, value=valor)
        c.font = Font(name="Arial", size=10)
        if i == 6 and total > 0:
            c.number_format = "0.0%"

    ws2["A13"] = "SOFTWARES / PLATAFORMAS MAIS FREQUENTES"
    ws2["A13"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")

    for col, title in [("A", "Software / Plataforma"), ("B", "Nº de Empresas")]:
        c = ws2[f"{col}14"]
        c.value = title
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COR_CAB)
        c.alignment = al_cen

    for i, (sw, cnt) in enumerate(sw_ord, 15):
        cor = COR_ALT if i % 2 == 0 else "FFFFFF"
        ws2[f"A{i}"] = sanitizar(sw)
        ws2[f"B{i}"] = cnt
        for col in ("A", "B"):
            ws2[f"{col}{i}"].fill = PatternFill("solid", start_color=cor)
            ws2[f"{col}{i}"].font = Font(name="Arial", size=10)
        ws2[f"B{i}"].alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 20

    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    print(f"\n✅ Planilha salva em: {caminho_saida}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Analisa PDFs de inventários GEE (v4 — seções 3.1/3.2/4.5/4.6).")
    parser.add_argument("--pasta",   "-p", default=PASTA_PDFS)
    parser.add_argument("--saida",   "-s", default=None)
    parser.add_argument("--paginas", "-n", type=int, default=MAX_PAGINAS)
    args = parser.parse_args()

    if args.saida is None:
        data_hoje = datetime.now().strftime("%Y-%m-%d")
        args.saida = str(Path(args.pasta) / f"resultados_gee_{data_hoje}.xlsx")

    print("=" * 65)
    print("  ANALISADOR DE INVENTÁRIOS GEE — SOFTWARES E PLATAFORMAS v4")
    print("  Seções analisadas: 3.1 · 3.2 · 4.5 · 4.6")
    print("=" * 65)

    resultados = analisar_pasta(args.pasta, args.paginas)
    if not resultados:
        print("Nenhum resultado para salvar.")
        return

    total      = len(resultados)
    com_sw     = sum(1 for r in resultados if r["usa_software"] == "SIM")
    erros      = sum(1 for r in resultados if r["usa_software"] == "ERRO")
    com_email  = sum(1 for r in resultados if r.get("email_responsavel", ""))
    com_secoes = sum(1 for r in resultados if r.get("secoes_encontradas") == "SIM")

    print(f"\n📊 Resumo:")
    print(f"   Total de PDFs:                   {total}")
    print(f"   Seções 3.1/3.2/4.5/4.6 extraídas:{com_secoes} ({100*com_secoes/total:.1f}%)")
    print(f"   Com software externo:            {com_sw} ({100*com_sw/total:.1f}%)")
    print(f"   Sem software externo:            {total - com_sw - erros}")
    print(f"   E-mails encontrados:             {com_email} ({100*com_email/total:.1f}%)")
    print(f"   Erros de leitura:                {erros}")

    gerar_excel(resultados, args.saida)


if __name__ == "__main__":
    main()
